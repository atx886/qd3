from selenium import webdriver
import time
from openpyxl import load_workbook, Workbook
from selenium.webdriver.chrome.options import Options
import os

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--disable-dev-shm-usage')
chromedriver = "/usr/bin/chromedriver"
os.environ["webdriver.chrome.driver"] = chromedriver

file = './Spider/2.xlsx'
max_a = load_workbook(file).active.max_row
print(max_a)
zheng = 0
jia = 0

import os
import re


def read_data(file):
    """读取文件中的内容"""
    with open(file, encoding='utf-8') as f:
        data = f.readlines()
        return data


def delete_file(file):
    """每次调用函数时，会删除文件的最后一行内容，并且将删除的内容返回"""
    file_data = read_data(file)
    # 删除列表中的最后一行，并且循环新列表中的内容，将新的内容写入文件中
    delete_data = file_data.pop()
    # # 判断原文件是否存在，如果存在，就删除,写文件时，会自动创建文件
    if os.path.exists(file):
        os.remove(file)
        for data in file_data:
            """将删除最后一行的列表数据写入文件中"""
            data = data.strip("\n")
            with open(file, mode="a", encoding="utf-8") as f:
                f.write(data)
                f.write("\n")
    return delete_data


def outip():
    file = r"IP代理池.txt"
    data = delete_file(file)
    t = re.findall(': \'(.+?)\'', data)
    # print(t[0])
    return t[0]


def writeexcle(t):
    wb = load_workbook(file)
    sheet = wb.active
    max_row = sheet.max_row - t

    row_max = 'a' + str(max_row)
    print(sheet)

    if max_row > 0:
        a = sheet[row_max].value
        print(a)
        return a
    else:
        print("空了")
        return None


# options = webdriver.FirefoxOptions()
# options.set_headless(True)
# options.add_argument("--headless")  # 设置火狐为headless无界面模式
# options.add_argument("--disable-gpu")
# d = webdriver.Firefox(options=options)

# d = webdriver.Firefox()
# # d = webdriver.Chrome()
# d.implicitly_wait(5)


def rw():
    time.sleep(1)


def dl(phone, d):
    d.get('https://www.chaojijishi.com/h5/#/pages/login/login?from=user')
    phh = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[1]/uni-input/div/input')

    rw()
    phh.send_keys(phone)

    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[2]/uni-input/div/input').send_keys(
        123456)
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[3]/uni-view[1]/uni-view').click()
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[5]/uni-view/uni-view').click()
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[4]').click()
    rw()
    d.get('https://www.chaojijishi.com/h5/#/pages/subpack1/set/user-id-card-data?type=1')
    return phone


def qd(d):
    d.get('https://www.chaojijishi.com/h5/#/pages/activityList/integral/integral')
    rw()

    dd = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[2]/uni-view[1]/uni-view[4]/uni-view[1]/uni-view[2]/uni-view')
    ms = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[2]/uni-view[1]/uni-view[4]/uni-view[1]/uni-view[2]/uni-view').text
    print(ms)
    a = 0
    while ms == "点击签到":
        dd.click()
        a += 1
        d.get('https://www.chaojijishi.com/h5/#/pages/activityList/integral/integral')
        rw()
        ms = d.find_element_by_xpath(
            '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[2]/uni-view[1]/uni-view[4]/uni-view[1]/uni-view[2]/uni-view').text
        if a > 5:
            return 0
    return 1


fal = []


def cs():
    a = max_a
    c = 0
    cheng = 0
    while a > 0:

        d = webdriver.Chrome(chrome_options=chrome_options, executable_path=chromedriver)
        # d = webdriver.Chrome()
        d.implicitly_wait(15)
        pho = ''
        try:
            pho = dl(writeexcle(c), d)
        except:
            c += 1
            sp = qd(d)
            cheng += sp
            fal.append(pho)
            a -= 1
            d.close()
            if c % 4 == 0:
                chrome_options.add_argument('--proxy-server=http://' + outip())
            continue
        c += 1
        sp = qd(d)
        cheng += sp
        if sp != 1:
            fal.append(pho)
        a -= 1
        d.close()
        if c % 4 == 0:
            chrome_options.add_argument('--proxy-server=http://' + outip())

    print('成功', cheng)
    print('shib', 128 - cheng)


cs()
